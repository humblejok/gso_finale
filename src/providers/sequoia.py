'''
Created on Oct 14, 2014

@author: sdejonckheere
'''
from django.db.models import Q
from openpyxl.reader.excel import load_workbook
from universe.models import PortfolioContainer, Attributes, RelatedCompany,\
    CompanyContainer
from utilities import external_content
import logging

LOGGER = logging.getLogger(__name__)

sequoia_dictionary = {'AR': Attributes.objects.get(active=True, identifier='SEQUOIA_CHARGE_TOP_AR'),
                      'IB': Attributes.objects.get(active=True, identifier='SEQUOIA_CHARGE_LOW_IB'),
                      'RM': Attributes.objects.get(active=True, identifier='SEQUOIA_CHARGE_LOW_RM'),
                      'MM': Attributes.objects.get(active=True, identifier='SEQUOIA_CHARGE_LOW_MM'),
                      'PM': Attributes.objects.get(active=True, identifier='SEQUOIA_CHARGE_LOW_PM'),
                      'MF': Attributes.objects.get(active=True, identifier='SEQUOIA_FEES_0_MF'),
                      'PF': Attributes.objects.get(active=True, identifier='SEQUOIA_FEES_1_PF'),
                      'OTHER': Attributes.objects.get(active=True, identifier='SEQUOIA_FEES_2_OF')
                      }
fees_set = ['SEQUOIA_FEES_0_MF', 'SEQUOIA_FEES_1_PF', 'SEQUOIA_FEES_2_OF']

def load_swm_map_file(path_to_file):
    workbook = load_workbook(path_to_file)
    sheet = workbook.get_sheet_by_name(name='SWM')
    row_index = 1
    sequoia_map = external_content.get_sequoia_map()
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    row_index = row_index + 1
    LOGGER.info("Using the following header: " + str(header))
    while row_index<=sheet.get_highest_row():
        LOGGER.info("Working on " + sheet.cell(row = row_index, column=1).value)
        current_name = sheet.cell(row = row_index, column=1).value
        portfolio = PortfolioContainer.objects.filter(Q(name=current_name) | Q(short_name=current_name))
        if portfolio.exists():
            portfolio = portfolio[0]
        else:
            portfolio = PortfolioContainer()
            portfolio.name = current_name
            portfolio.short_name = current_name
        portfolio.inception_date = sheet.cell(row = row_index, column=2).value
        portfolio.currency = Attributes.objects.get(short_name__iexact=sheet.cell(row = row_index, column=3).value, type='currency', active=True)
        portfolio.status = Attributes.objects.get(type='status', identifier='STATUS_ACTIVE', active=True)
        portfolio.type = Attributes.objects.get(type='container_type', identifier='CONT_PORTFOLIO', active=True)
        portfolio.save()
        associated_bank = portfolio.associated_companies.filter(Q(company__name__iexact=sheet.cell(row = row_index, column=7).value) | Q(company__short_name__iexact=sheet.cell(row = row_index, column=7).value), role__identifier='SCR_BANK')
        if not associated_bank.exists():
            current_bank = portfolio.associated_companies.filter(role__identifier='SCR_BANK')
            if current_bank.exists():
                current_bank = current_bank[0]
                portfolio.associated_companies.remove(current_bank)
            associated_bank = CompanyContainer.objects.filter(Q(name__iexact=sheet.cell(row = row_index, column=7).value) | Q(short_name__iexact=sheet.cell(row = row_index, column=7).value))
            if associated_bank.exists():
                associated_bank = associated_bank[0]
            else:
                associated_bank = CompanyContainer()
                associated_bank.name = sheet.cell(row = row_index, column=7).value
                associated_bank.short_name = sheet.cell(row = row_index, column=7).value
                associated_bank.status = Attributes.objects.get(type='status', identifier='STATUS_TO_BE_VALIDATED', active=True)
                associated_bank.type = Attributes.objects.get(type='container_type', identifier='CONT_COMPANY', active=True)
                associated_bank.save()
            current_bank = RelatedCompany()
            current_bank.role = Attributes.objects.get(identifier='SCR_BANK', active=True)
            current_bank.company = associated_bank
            current_bank.save()
            portfolio.associated_companies.add(current_bank)
        else:
            associated_bank = associated_bank[0].company
        portfolio_id = str(portfolio.id)
        sequoia_map[portfolio_id] = external_content.create_sequoia_map_entry(portfolio)
        strategy_profile = Attributes.objects.get(Q(short_name__iexact=sheet.cell(row = row_index, column=4).value) | Q(name__iexact=sheet.cell(row = row_index, column=4).value), Q(type='sequoia_strategy'), Q(active=True))
        risk_profile = Attributes.objects.get(Q(short_name__iexact=sheet.cell(row = row_index, column=5).value) | Q(name__iexact=sheet.cell(row = row_index, column=5).value), Q(type='sequoia_risk'), Q(active=True))
        jurisdiction = Attributes.objects.filter(Q(short_name__iexact=sheet.cell(row = row_index, column=6).value) | Q(name__iexact=sheet.cell(row = row_index, column=6).value), Q(type='country_iso2'), Q(active=True))
        if jurisdiction.exists():
            jurisdiction = jurisdiction[0]
        else:
            LOGGER.warn("\tJurisdiction not found [" + sheet.cell(row = row_index, column=6).value + "]")
            jurisdiction = None
        LOGGER.info("\tUsing key:" + portfolio_id)
        sequoia_map[portfolio_id]['strategy_profile'] = strategy_profile.identifier
        sequoia_map[portfolio_id]['risk_profile'] = risk_profile.identifier
        sequoia_map[portfolio_id]['jurisdiction'] = jurisdiction.identifier if jurisdiction!=None else None
        sequoia_map[portfolio_id]['bank'] = associated_bank.name
        # Cleaning fees setup
        for fee in fees_set:
            sequoia_map[portfolio_id][fee] = {}
        # Parsing ratios
        structure_ratio = sheet.cell(row = row_index, column=8).value
        sequoia_map[portfolio_id]['structure_ratio'] = structure_ratio
        for index in xrange(10,sheet.get_highest_column() - 3,3):
            if sheet.cell(row = row_index, column=index).value!=None and sheet.cell(row = row_index, column=index).value!='':
                key = header[index-1].upper().split(' ')
                charge = sequoia_dictionary[key[0]].identifier
                fee = sequoia_dictionary[key[1]].identifier
                if not sequoia_map[portfolio_id][fee].has_key(charge):
                    sequoia_map[portfolio_id][fee][charge] = []
                bud = Attributes.objects.filter(active=True, type='sequoia_bud', short_name=sheet.cell(row = row_index, column=index).value)
                if bud.exists():
                    bud = bud[0].identifier
                    sequoia_map[portfolio_id][fee][charge].append({'rate':sheet.cell(row = row_index, column=index + 1).value * 100.0, 'bud': bud})
                else:

                    LOGGER.warn("\tBUD not found [" + str(sheet.cell(row = row_index, column=index).value) + "]")
        row_index = row_index + 1
    external_content.set_sequoia_map(sequoia_map)
    
    