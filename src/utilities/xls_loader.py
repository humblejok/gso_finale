'''
Created on May 24, 2014

@author: sdejonckheere
'''
from openpyxl.reader.excel import load_workbook
import logging

LOGGER = logging.getLogger(__name__)

def load_xls(sheet_name, xlsx_file, group_by=None):
    None
    
def load_csv(csv_file, group_by=None):
    None

def load_xlsx(sheet_name, xlsx_file, group_by=None):
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    row_index = 0
    # Reading header
    header = []
    for column_index in range(0, sheet.get_highest_column()):
        value = sheet.cell(row=row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    
    row_index += 1
    if group_by!=None:
        all_data = {}
    else:
        all_data = []
    keys = None
    while row_index<sheet.get_highest_row():
        data_set = {}
        if group_by!=None:
            keys = {}
            for identifier in group_by:
                keys[identifier] = sheet.cell(row=row_index, column=header.index(identifier)).value
        for index in range(0, len(header)):
            data_set[header[index]] = sheet.cell(row=row_index, column=index).value
        if keys!=None:
            working_data = all_data
            for identifier in group_by:
                if not working_data.has_key(keys[identifier]):
                    if group_by[len(group_by)-1]==identifier:
                        working_data[keys[identifier]] = []
                    else:
                        working_data[keys[identifier]] = {}
                working_data = working_data[keys[identifier]]
        if keys!=None:
            working_data = all_data
            for identifier in group_by:
                working_data = working_data[keys[identifier]]
            working_data.append(data_set)
        else:
            all_data.append(data_set)
        row_index += 1
    return all_data
        