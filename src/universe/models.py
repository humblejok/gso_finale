import csv
from datetime import datetime as dt
import datetime
import logging
import os
import traceback

from django.contrib.auth.models import User
from django.contrib.contenttypes import generic
from django.contrib.contenttypes.models import ContentType
from django.db import models
from django.db.models import Q
from django.db.models.fields import FieldDoesNotExist, DateTimeField
from django.template import loader
from django.template.context import Context
from openpyxl.reader.excel import load_workbook
from seq_common.utils import classes, dates
import xlrd
from xlrd.xldate import xldate_as_tuple

from finale.settings import RESOURCES_MAIN_PATH, STATICS_PATH, \
    STATICS_GLOBAL_PATH
import universe
from utilities import computing
from utilities.track_content import set_track_content, get_track_content


LOGGER = logging.getLogger(__name__)

def setup():
    setup_attributes()
    setup_labels()
    populate_bloomberg_fields(os.path.join(RESOURCES_MAIN_PATH,'fields.csv'))
    populate_model_from_xlsx('universe.models.MenuEntries', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('universe.models.BloombergDataContainerMapping', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('universe.models.BloombergTrackContainerMapping', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('universe.models.CompanyContainer', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('universe.models.AccountContainer', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_model_from_xlsx('universe.models.PersonContainer', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_attributes()

def setup_menus():
    MenuEntries.objects.all().delete()
    populate_model_from_xlsx('universe.models.MenuEntries', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))

def setup_labels():
    FieldLabel.objects.all().delete()
    populate_labels_from_xlsx('universe.models.FieldLabel', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))

def setup_attributes():
    populate_attributes_from_xlsx('universe.models.Attributes', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_attributes_from_xlsx('universe.models.Dictionary', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_attributes()

def generate_attributes():
    all_types = Attributes.objects.all().order_by('type').distinct('type')
    context = Context({"selection": all_types})
    template = loader.get_template('rendition/attributes_list_option_renderer.html')
    rendition = template.render(context)
    # TODO Implement multi-langage
    outfile = os.path.join(STATICS_PATH, 'all_types_option_en.html')
    with open(outfile,'w') as o:
        o.write(rendition.encode('utf-8'))
    template = loader.get_template('rendition/attributes_list_select_renderer.html')
    rendition = template.render(context)
    # TODO Implement multi-langage
    outfile = os.path.join(STATICS_GLOBAL_PATH, 'all_types_select_en.html')
    with open(outfile,'w') as o:
        o.write(rendition.encode('utf-8'))
    for a_type in all_types:
        all_elements = Attributes.objects.filter(type=a_type.type, active=True)
        context = Context({"selection": all_elements})
        template = loader.get_template('rendition/attributes_option_renderer.html')
        rendition = template.render(context)
        # TODO Implement multi-langage
        outfile = os.path.join(STATICS_PATH, a_type.type + '_en.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))
        template = loader.get_template('rendition/attributes_select_renderer.html')
        rendition = template.render(context)
        # TODO Implement multi-langage
        outfile = os.path.join(STATICS_GLOBAL_PATH, a_type.type + '_select_en.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))
        template = loader.get_template('rendition/attributes_list_elements_renderer.html')
        rendition = template.render(context)
        # TODO Implement multi-langage
        outfile = os.path.join(STATICS_GLOBAL_PATH, a_type.type + '_list_elements_en.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))

def populate_attributes_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        if model.objects.filter(identifier=sheet.cell(row = row_index, column=1).value).exists():
            instance = model.objects.get(identifier=sheet.cell(row = row_index, column=1).value)
        else:
            instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i+1).value
            setattr(instance, header[i], value)
        instance.save()
        row_index += 1
        
def populate_labels_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        if model.objects.filter(identifier=sheet.cell(row = row_index, column=1).value).exists():
            instance = model.objects.get(identifier=sheet.cell(row = row_index, column=1).value, langage=sheet.cell(row = row_index, column=2))
        else:
            instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i+1).value
            setattr(instance, header[i], value)
        instance.save()
        row_index += 1

def populate_model_from_xlsx(model_name, xlsx_file):
    LOGGER.info("Loading data in " + model_name)
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i+1).value
            field_info = Attributes()
            field_info.short_name = header[i]
            field_info.name = header[i]
            instance.set_attribute('excel', field_info, value)
        instance.save()
        row_index += 1
        
def populate_bloomberg_fields(csv_file):
    BloombergField.objects.all().delete()
    with open(csv_file) as bb_fields:
        reader = csv.reader(bb_fields,delimiter=',', quotechar='"', skipinitialspace=True)
        first = True
        for row in reader:
            if first:
                header = row
                LOGGER.info('Using header:' + str(header))
                first = False
            else:
                field = BloombergField()
                field.identifier = row[header.index('Field ID')]
                field.code = row[header.index('Field Mnemonic')]
                field.name = row[header.index('Field Mnemonic')]
                field.short_name = row[header.index('Field Mnemonic')]
                field.descrition = row[header.index('Description')]
                field.dl_category = row[header.index('Data License Category')]
                field.category = row[header.index('Category')]
                field.definition = row[header.index('Definition')]
                field.commodity = row[header.index('Comdty')].strip()!=''
                field.equity = row[header.index('Equity')].strip()!=''
                field.municipal_bond = row[header.index('Muni')].strip()!=''
                field.preferred_share = row[header.index('Pfd')].strip()!=''
                field.money_market = row[header.index('M-Mkt')].strip()!=''
                field.government_bond = row[header.index('Govt')].strip()!=''
                field.corporate_bond = row[header.index('Corp')].strip()!=''
                field.index = row[header.index('Index')].strip()!=''
                field.currency = row[header.index('Curncy')].strip()!=''
                field.mortgage = row[header.index('Mtge')].strip()!=''
                field.width = row[header.index('Standard Width')]
                field.decimals = row[header.index('Standard Decimal Places')]
                field.type = row[header.index('Field Type')]
                field.get_fundamentals = row[header.index('Getfundamentals')].strip()!=''
                field.get_history = row[header.index('Gethistory')].strip()!=''
                field.get_company = row[header.index('Getcompany')].strip()!=''
                field.dl_2nd_category = row[header.index('Data License Category 2')]
                field.save()

def populate_perf(container, source_track, reference=None):
    LOGGER.info('Computing and saving ' + source_track.frequency.name + ' performances track for ' + container.name)
    perf = Attributes.objects.get(identifier='NUM_TYPE_PERF', active=True)
    computing_company = CompanyContainer.objects.get(name='FinaLE Engine')
    try:
        if reference==None:
            track = TrackContainer.objects.get(
                                effective_container_id=container.id,
                                type__id=perf.id,
                                quality__id=source_track.quality.id,
                                source__id=computing_company.id,
                                frequency__id=source_track.frequency.id,
                                status__id=source_track.status.id)
        else:
            track = TrackContainer.objects.get(
                                effective_container_id=container.id,
                                type__id=perf.id,
                                quality__id=source_track.quality.id,
                                source__id=computing_company.id,
                                frequency__id=source_track.frequency.id,
                                status__id=source_track.status.id,
                                frequency_reference=reference)
        LOGGER.info("Track already exists")
    except:
        track = TrackContainer()
        track.effective_container = SecurityContainer.objects.get(id=container.id)
        track.type = perf
        track.quality = source_track.quality
        track.source = computing_company
        track.status = source_track.status
        track.frequency = source_track.frequency
        track.frequency_reference = reference
        track.save()
    computer = computing.get_tracks_computer()
    LOGGER.info(source_track.frequency.name + ' performances computation starts for ' + container.name)
    track_content = get_track_content(source_track, True)
    performances = computer.compute_performances([value['value'] for value in track_content])
    LOGGER.info(source_track.frequency.name + ' performances computation ends for ' + container.name + '. Got ' + str(len(performances)) + ' elements!')
    performances_content = []
    for index in range(1,len(performances)):
        performances_content.append({'date':track_content[index]['date'],'value':performances[index]})
    set_track_content(track, performances_content, True)
    LOGGER.info(source_track.frequency.name + ' performances track is now in database for ' + container.name)

def populate_weekly_track_from_track(container, source_track):
    LOGGER.info('Computing weekly prices track for ' + container.name)
    track_content = get_track_content(source_track, True)
    reference_days = [Attributes.objects.get(identifier=day) for day in ['DT_REF_MONDAY','DT_REF_TUESDAY','DT_REF_WEDNESDAY','DT_REF_THURSDAY','DT_REF_FRIDAY','DT_REF_SATURDAY','DT_REF_SUNDAY']]
    weekly = Attributes.objects.get(identifier='FREQ_WEEKLY', active=True)
    computing_company = CompanyContainer.objects.get(name='FinaLE Engine')
    for day in reference_days[0:5]:
        LOGGER.info('Working on day:' + day.name)
        try:
            track = TrackContainer.objects.get(
                                effective_container_id=container.id,
                                type__id=source_track.type.id,
                                quality__id=source_track.quality.id,
                                source__id=computing_company.id,
                                frequency__id=weekly.id,
                                frequency_reference=day,
                                status__id=source_track.status.id)
            LOGGER.info("Track already exists")
        except:
            track = TrackContainer()
            track.effective_container = SecurityContainer.objects.get(id=container.id)
            track.type = source_track.type
            track.quality = source_track.quality
            track.source = computing_company
            track.status = source_track.status
            track.frequency = weekly
            track.frequency_reference = day
            track.save()
        
        started = False
        
        all_values = []
        
        previous = None
        next_date = None
        for token in track_content:
            if started:
                if token['date']==next_date:
                    all_values.append(token)
                    next_date = token['date'] + datetime.timedelta(days=7)
                elif token['date']>next_date:
                    all_values.append({'date': next_date, 'value':previous['value']})
                    next_date = next_date + datetime.timedelta(days=7)
                previous = token
            else:
                if reference_days[token['date'].weekday()].id==day.id:
                    started = True
                    all_values.append(token)
                    previous = token
                    next_date = token['date'] + datetime.timedelta(days=7)
        set_track_content(track, all_values, True)
        populate_perf(container, track, day)


def populate_monthly_track_from_track(container, source_track):
    LOGGER.info('Computing monthly prices track for ' + container.name)
    monthly = Attributes.objects.get(identifier='FREQ_MONTHLY', active=True)
    computing_company = CompanyContainer.objects.get(name='FinaLE Engine')
    try:
        track = TrackContainer.objects.get(
                            effective_container_id=container.id,
                            type__id=source_track.type.id,
                            quality__id=source_track.quality.id,
                            source__id=computing_company.id,
                            frequency__id=monthly.id,
                            status__id=source_track.status.id)
        LOGGER.info("Track already exists")
    except:
        track = TrackContainer()
        track.effective_container = SecurityContainer.objects.get(id=container.id)
        track.type = source_track.type
        track.quality = source_track.quality
        track.source = computing_company
        track.status = source_track.status
        track.frequency = monthly
        track.frequency_reference = None
        track.save()
    track_tokens = get_track_content(source_track, True)
    LOGGER.info('Working on ' + str(len(track_tokens)) + ' elements!' )
    all_values = []
    previous_token = None
    for token in track_tokens:
        if previous_token!=None:
            current_eom = datetime.datetime.combine(dates.GetEndOfMonth(previous_token['date']), dt.min.time())
            if (token['date']-current_eom).days!=0 and previous_token['date'].month!=token['date'].month and previous_token['date']!=dates.GetEndOfMonth(previous_token['date']):
                all_values.append({'date': current_eom, 'value':token['value'] if token['date']==current_eom else previous_token['value']})
        else:
            current_eom = datetime.datetime.combine(dates.GetEndOfMonth(token['date']), dt.min.time())
            if current_eom!=token['date']:
                all_values.append({'date': datetime.datetime.combine(dates.GetEndOfMonth(dates.AddMonth(token['date'],-1)), dt.min.time()), 'value':token['value'] })
        previous_token = token
    LOGGER.info('Finished monthly prices track computation for ' + container.name)
    set_track_content(track, all_values, True)
    populate_perf(container, track)
    
def populate_track_from_lyxor(lyxor_file):
    # Excel input
    workbook = xlrd.open_workbook(lyxor_file)
    sheet = workbook.sheet_by_name('Weekly NAV')
    # Lyxor default
    lyxor_company = CompanyContainer.objects.get(name='Lyxor Asset Management')
    nav_value = Attributes.objects.get(identifier='NUM_TYPE_NAV', active=True)
    final_status = Attributes.objects.get(identifier='NUM_STATUS_FINAL', active=True)
    official_type = Attributes.objects.get(identifier='PRICE_TYPE_OFFICIAL', active=True)
    weekly = Attributes.objects.get(identifier='FREQ_WEEKLY', active=True)
    tuesday = Attributes.objects.get(identifier='DT_REF_THURSDAY', active=True)
    row_index = 9
    isin_cache = {}
    isin_content = {} 
    while row_index<sheet.nrows:
        full_date = datetime.datetime(*xldate_as_tuple(sheet.cell_value(row_index, 1), workbook.datemode))
        simple_date = datetime.date(full_date.year, full_date.month, full_date.day)
        LOGGER.info("DATE:" + str(simple_date))
        for col_index in range(2,sheet.ncols):
            if sheet.cell_type(row_index,col_index)!=xlrd.XL_CELL_EMPTY and sheet.cell_value(row_index,col_index)!=None and sheet.cell_value(row_index,col_index)!='':
                isin = sheet.cell_value(5,col_index)
                if not isin_cache.has_key(isin):
                    targets = SecurityContainer.objects.filter(aliases__alias_type__name='ISIN', aliases__alias_value=isin)
                    isin_cache[isin] = targets
                else:
                    targets = isin_cache[isin]
                LOGGER.info("Updating " + str(len(targets)) + " entities with ISIN like " + str(isin));
                if not isin_content.has_key(isin):
                    isin_content[isin] = []
                isin_content[isin].append({'date': dt.combine(simple_date, dt.min.time()), 'value':float(sheet.cell_value(row_index,col_index))})
        row_index += 1
    for isin in isin_cache.keys():
        for target in isin_cache[isin]:
            try:
                track = TrackContainer.objects.get(effective_container_id=target.id, frequency_reference=tuesday, type__id=nav_value.id, quality__id=official_type.id,source__id=lyxor_company.id, frequency__id=weekly.id, status__id=final_status.id)
                LOGGER.info("Track already exists")
            except:
                track = TrackContainer()
                track.effective_container = SecurityContainer.objects.get(id=target.id)
                track.type = nav_value
                track.quality = official_type
                track.source = lyxor_company
                track.status = final_status
                track.frequency = weekly
                track.frequency_reference = tuesday
                track.save()
            set_track_content(track, isin_content[isin], True)
            populate_perf(track.effective_container, track)
            populate_monthly_track_from_track(track.effective_container, track)

def populate_tracks_from_bloomberg_protobuf(data, update=False):
    LOGGER.info('Importing historical data from Bloomberg')
    official_type = Attributes.objects.get(identifier='PRICE_TYPE_OFFICIAL', active=True)
    daily = Attributes.objects.get(identifier='FREQ_DAILY', active=True)
    bloomberg_company = CompanyContainer.objects.get(name='Bloomberg LP')
    final_status = Attributes.objects.get(identifier='NUM_STATUS_FINAL', active=True)
    not_found = []
    cache = {}
    all_values = {}
    for row in data.rows:
        if row.errorCode==0 and not not_found.__contains__(row.ticker):
            work_date = datetime.datetime.fromtimestamp(row.date/1000.0)
            work_date = datetime.date(work_date.year, work_date.month, work_date.day)
            target = None
            with_isin = SecurityContainer.objects.filter(aliases__alias_type__name='ISIN', aliases__alias_value=row.ticker)
            with_bloomberg = SecurityContainer.objects.filter(aliases__alias_type__name='BLOOMBERG', aliases__alias_value=row.ticker)
            if cache.has_key(row.ticker):
                target = cache[row.ticker]
            elif with_isin.exists():
                target = with_isin[0]
                cache[row.ticker] = target
            elif with_bloomberg.exists():
                target = with_bloomberg[0]
                cache[row.ticker] = target
            if target==None:
                LOGGER.warn('Could not find a security container with ISIN or BLOOMBERG code equals to [' + str(row.ticker) + ']')
                not_found.append(row.ticker)
            else:
                if not all_values.has_key(target.id):
                    all_values[target.id] = {}
                if not all_values[target.id].has_key(row.field):
                    all_values[target.id][row.field] = []
                all_values[target.id][row.field].append({'date': dt.combine(work_date, dt.min.time()), 'value': row.valueDouble})
    for key in all_values.keys():
        for field in all_values[key].keys():
            current_type = BloombergTrackContainerMapping.objects.get(short_name__code=field).name
            try:
                track = TrackContainer.objects.get(effective_container_id=key,type__id=current_type.id,quality__id=official_type.id,source__id=bloomberg_company.id, frequency__id=daily.id, status__id=final_status.id)
                LOGGER.info("Track already exists")
            except:
                track = TrackContainer()
                track.effective_container = SecurityContainer.objects.get(id=key)
                track.type = current_type
                track.quality = official_type
                track.source = bloomberg_company
                track.status = final_status
                track.frequency = daily
                track.frequency_reference = None
                track.save()
            set_track_content(track, all_values[key][field], not update)
    LOGGER.info('Historical tracks imported from Bloomberg')
    nav_value = Attributes.objects.get(identifier='NUM_TYPE_NAV', active=True)
    for key in all_values.keys():
        # Not always correct, check about NUM_TYPE_NAV instead
        if all_values[key].has_key('PX_LAST'):
            track = TrackContainer.objects.get(effective_container_id=key,type__id=nav_value.id,quality__id=official_type.id,source__id=bloomberg_company.id, frequency__id=daily.id, status__id=final_status.id)
            if len(all_values[key])>0:
                populate_perf(track.effective_container, track)
                populate_monthly_track_from_track(track.effective_container, track)
                populate_weekly_track_from_track(track.effective_container, track)
            else:
                LOGGER.warn("Empty track for " + key + " no computation will be executed!")

def populate_security_from_lyxor(lyxor_file, clean=False):
    universe = Universe.objects.filter(short_name='LYXOR')
    if universe.exists():
        universe = universe[0]
        universe.members.clear()
        universe.save()
    else:
        universe = Universe()
        universe.name = 'Lyxor Universe'
        universe.short_name = 'LYXOR'
        universe.type = Attributes.objects.get(identifier='CONT_UNIVERSE')
        universe.inception_date = datetime.date.today()
        universe.closed_date = None
        universe.status = Attributes.objects.get(identifier='STATUS_ACTIVE')
        universe.public = True
        universe.owner = User.objects.get(id=1)
        universe.description = "This universe contains all Lyxor B ETFs and trackers."
        universe.save()
    if clean:
        SecurityContainer.objects.all().delete()
    
    lyxor_company = CompanyContainer.objects.get(name='Lyxor Asset Management')
    
    data_provider = Attributes.objects.get(identifier='SCR_DP', active=True)
    lyxor_provider = RelatedCompany.objects.filter(company=lyxor_company, role=data_provider)
    if lyxor_provider.exists():
        lyxor_provider = lyxor_provider[0]
    else:
        lyxor_provider = RelatedCompany()
        lyxor_provider.company = lyxor_company
        lyxor_provider.role = data_provider
        lyxor_provider.save()
    
    usd_currency = Attributes.objects.get(identifier='CURR_USD', active=True)
    weekly = Attributes.objects.get(identifier='FREQ_WEEKLY', active=True)
    tuesday = Attributes.objects.get(identifier='DT_REF_THURSDAY', active=True)
    security_type = Attributes.objects.get(identifier='CONT_FUND')
    fund_type = Attributes.objects.get(identifier='SECTYP_FUND')
    workbook = xlrd.open_workbook(lyxor_file)
    sheet = workbook.sheet_by_name('Lyxor Accounts')
    row_index = 6
    # Reading header
    HEADER = []
    for column_index in range(1, sheet.ncols):
        value = sheet.cell_value(row_index, column_index)
        HEADER.append(value if value!='' else HEADER[-1])
    LOGGER.info('Using header:' + str(HEADER))
    row_index += 1
    while row_index<sheet.nrows:
        value = sheet.cell_value(row_index,2)
        if not sheet.cell_type(row_index,2)==xlrd.XL_CELL_EMPTY and value.strip()!='':
            update = SecurityContainer.objects.filter(aliases__alias_type__name='ISIN', aliases__alias_value=value)
            if update.exists():
                LOGGER.info('Found ' + str(len(update)) + ' securities with identifier [' + str(value) + ']')
                update = [SecurityContainer.objects.get(id=security.id) for security in update]
                [security.clean() for security in update]
                [security.save() for security in update]
            else:
                update = [SecurityContainer.create()]
                [security.save() for security in update]
                LOGGER.info('Creating new entry ' + SecurityContainer.__name__)
            for i in range(1,sheet.ncols):
                value = sheet.cell_value(row_index,i)
                if sheet.cell_type(row_index,i)==xlrd.XL_CELL_DATE:
                    value = datetime.date(*xldate_as_tuple(value, workbook.datemode)[:3])
                field_info = Attributes.objects.filter(type='lyxor_field', name=HEADER[i-1])
                if field_info.exists():
                    [security.set_attribute('lyxor', field_info[0], value) for security in update]
                else:
                    LOGGER.debug("Cannot find matching field for " + HEADER[i-1])
            [setattr(security,'type', security_type) for security in update]
            [setattr(security,'currency', usd_currency) for security in update]
            # Update currency if needed
            for security in update:
                currency = security.get_currency_in_name()
                if currency!=None:
                    security.currency = currency
            [setattr(security,'security_type', fund_type) for security in update]
            [setattr(security,'frequency', weekly) for security in update]
            [setattr(security,'frequency_reference', tuesday) for security in update]
            [security.finalize() for security in update]
            [security.associated_companies.add(lyxor_provider) for security in update]
            [security.save() for security in update]
            [LOGGER.info("Aliases:" + str(len(security.aliases.all()))) for security in update]
            [universe.members.add(security) for security in update]
        row_index += 1


class CoreModel(models.Model):

    def get_editable_fields(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()!='ManyToManyField':
                values.append(field)
        return values
        
    def get_associable_field(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()=='ManyToManyField':
                values.append(field)
        return values        

    @staticmethod
    def get_fields():
        return []
    
    def get_identifier(self):
        return 'name'
    
    def list_values(self):
        values = []
        for field in self.get_fields():
            LOGGER.debug(self.__class__.__name__ + ' * ' + field)
            if field in self._meta.get_all_field_names():
                if self._meta.get_field(field).get_internal_type()=='ManyToManyField' and getattr(self,field)!=None:
                    values.append(str([e.list_values() for e in list(getattr(self,field).all())]))
                elif self._meta.get_field(field).get_internal_type()=='ForeignKey' and getattr(self,field)!=None:
                    values.append(getattr(self,field).get_value())
                else:
                    values.append(str(getattr(self,field)))
            else:
                # Generic foreign key
                values.append(getattr(self,field).get_value())
        return values
    
    def get_value(self):
        if self.get_identifier()!=None:
            return getattr(self, self.get_identifier())
        else:
            return None
        
    def __unicode__(self):
        return unicode(self.get_value())
    
    def set_attribute(self, source, field_info, string_value):
        try:
            if string_value!='' and string_value!=None:
                if self._meta.get_field(field_info.short_name).get_internal_type()=='ManyToManyField':
                    if not self.many_fields.has_key(field_info.short_name):
                        self.many_fields[field_info.short_name] = []
                    foreign = self._meta.get_field(field_info.short_name).rel.to
                    foreign_element = foreign.retrieve_or_create(self, source, field_info.name, string_value)
                    if foreign_element!=None:                
                        self.many_fields[field_info.short_name].append(foreign_element)
                elif self._meta.get_field(field_info.short_name).get_internal_type()=='DateField' or self._meta.get_field(field_info.short_name).get_internal_type()=='DateTimeField':
                    try:
                        dt = datetime.datetime.strptime(string_value,'%m/%d/%Y' if source!='web' else '%y-%m-%d')
                        if self._meta.get_field(field_info.short_name).get_internal_type()=='DateField':
                            dt = datetime.date(dt.year, dt.month, dt.day)
                    except:
                        dt = string_value # This is not a String???
                    setattr(self, field_info.short_name, dt)
                elif self._meta.get_field(field_info.short_name).get_internal_type()=='ForeignKey':
                    linked_to = self._meta.get_field(field_info.short_name).rel.limit_choices_to
                    foreign = self._meta.get_field(field_info.short_name).rel.to
                    filtering_by_name = dict(linked_to)
                    filtering_by_name['name'] = string_value
                    by_name = foreign.objects.filter(**filtering_by_name)
                    filtering_by_short = dict(linked_to)
                    filtering_by_short['short_name'] = string_value
                    by_short = foreign.objects.filter(**filtering_by_short)
                    if foreign==universe.models.Attributes:
                        filtering_by_identifier = dict(linked_to)
                        filtering_by_identifier['identifier'] = string_value
                        by_identifier = foreign.objects.filter(**filtering_by_identifier)
                    else:
                        by_identifier = None
                    if by_name.exists():
                        setattr(self, field_info.short_name, by_name[0])
                    elif by_short.exists():
                        setattr(self, field_info.short_name, by_short[0])
                    elif by_identifier!=None and by_identifier.exists():
                        setattr(self, field_info.short_name, by_identifier[0])
                    else:
                        dict_entry = Dictionary.objects.filter(name=linked_to['type'], auto_create=True)
                        if dict_entry.exists():
                            LOGGER.info('Creating new attribute for ' + linked_to['type'] + ' with value ' + string_value)
                            new_attribute = Attributes()
                            new_attribute.active = True
                            new_attribute.identifier = dict_entry[0].identifier + str(string_value.upper()).replace(' ', '_')
                            new_attribute.name = string_value
                            new_attribute.short_name = string_value[0:32]
                            new_attribute.type = linked_to['type']
                            new_attribute.save()
                            setattr(self, field_info.short_name, new_attribute)
                        else:
                            LOGGER.warn('Cannot find foreign key instance on ' + str(self) + '.' + field_info.short_name + ' for value [' + string_value + '] and relation ' + str(linked_to))
                else:
                    setattr(self, field_info.short_name, string_value)
        except FieldDoesNotExist:
            traceback.print_exc()
            LOGGER.error("Wrong security type for " + self.name + ", please check your settings...")
    
    class Meta:
        ordering = ['id']

class FieldLabel(CoreModel):
    identifier = models.CharField(max_length=512)
    langage = models.CharField(max_length=3)
    field_label = models.CharField(max_length=1024)
    
    @staticmethod
    def get_fields():
        return ['identifier','langage','field_label']
    
    def get_identifier(self):
        return 'identifier'
    
class Attributes(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    type = models.CharField(max_length=64)
    active = models.BooleanField()
    
    @staticmethod
    def get_fields():
        return ['identifier','name','short_name','type','active']
    
    def get_short_json(self):
        return {'id': self.id, 'name': self.name, 'short_name': self.short_name, 'identifier': self.identifier }

    class Meta:
        ordering = ['name']
        

class MenuEntries(CoreModel):
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    menu_target = models.ForeignKey(Attributes, limit_choices_to={'type':'container_menu_target'}, related_name='container_menu_target_rel', null=True)
    menu_type = models.CharField(max_length=128)
    container_type = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='container_type_menu_rel', null=True)
    language = models.CharField(max_length=3)
    data_target = models.CharField(max_length=128)
    action_type = models.CharField(max_length=128, null=True, blank=True)
    action_label = models.CharField(max_length=128)
    
    @staticmethod
    def get_fields():
        return ['menu_target','menu_type','container_type','language','data_target','action_type','action_label']

    class Meta:
        ordering = ['menu_target']
    
class BloombergField(CoreModel):
    identifier = models.CharField(max_length=128)
    code = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=128)
    descrition = models.TextField()
    dl_category = models.CharField(max_length=128, null=True, blank=True)
    category = models.CharField(max_length=128, null=True, blank=True)
    definition = models.TextField()
    commodity = models.BooleanField(default=False)
    equity = models.BooleanField(default=False)
    municipal_bond = models.BooleanField(default=False)
    preferred_share = models.BooleanField(default=False)
    money_market = models.BooleanField(default=False)
    government_bond = models.BooleanField(default=False)
    corporate_bond = models.BooleanField(default=False)
    index = models.BooleanField(default=False)
    currency = models.BooleanField(default=False)
    mortgage = models.BooleanField(default=False)
    width = models.IntegerField(null=True)
    decimals = models.IntegerField(null=True)
    type = models.CharField(max_length=64)
    get_fundamentals = models.BooleanField(default=False)
    get_history = models.BooleanField(default=False)
    get_company = models.BooleanField(default=False)
    dl_2nd_category = models.CharField(max_length=128, null=True, blank=True)
    
class Dictionary(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    auto_create = models.BooleanField()
    
    @staticmethod
    def get_fields():
        return ['identifier','name','auto_create']
    
    class Meta:
        ordering = ['name']
        
class Address(CoreModel):
    address_type = models.ForeignKey(Attributes, limit_choices_to={'type':'address_type'}, related_name='address_type_rel', null=True)
    line_1 = models.CharField(max_length=128, null=True)
    line_2 = models.CharField(max_length=128, null=True)
    zip_code = models.CharField(max_length=16, null=True)
    city = models.CharField(max_length=128, null=True)
    country = models.ForeignKey(Attributes, limit_choices_to={'type':'country_iso2'}, related_name='address_country_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['address_type','line_1','line_2','zip_code','city','country']
    
    @staticmethod
    def get_filtering_field():
        return "address_type"
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['address_type.name', 'line_1','line_2','zip_code','city', 'country.name']
        elif rendition_width=='small':
            return ['address_type.name', 'city', 'country.name']

class Email(CoreModel):
    address_type = models.ForeignKey(Attributes, limit_choices_to={'type':'email_type'}, related_name='email_type_rel', null=True)
    email_address = models.EmailField()

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['address_type','email_address']
    
    @staticmethod
    def get_filtering_field():
        return "address_type"
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['address_type.name', 'email_address']
        elif rendition_width=='small':
            return ['address_type.name', 'email_address']
    
class Phone(CoreModel):
    line_type = models.ForeignKey(Attributes, limit_choices_to={'type':'phone_type'}, related_name='phone_type_rel', null=True)
    phone_number = models.TextField(max_length=32)

    @staticmethod
    def get_filtering_field():
        return "line_type"

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['line_type','phone_number']
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['line_type.name', 'phone_number']
        elif rendition_width=='small':
            return ['line_type.name', 'phone_number']
    
class Alias(CoreModel):
    alias_type = models.ForeignKey(Attributes, limit_choices_to={'type':'alias_type'}, related_name='alias_type_rel', null=True)
    alias_value = models.CharField(max_length=512)
    alias_additional = models.CharField(max_length=512)
    
    def get_identifier(self):
        return 'id'
    
    @staticmethod
    def get_fields():
        return ['alias_type','alias_value','alias_additional']
    
    @staticmethod
    def get_querying_fields():
        return ['alias_value','alias_additional']

    @staticmethod
    def get_filtering_field():
        return "alias_type"
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['alias_type.name','alias_value','alias_additional']
        elif rendition_width=='small':
            return ['alias_type.name','alias_value']
        
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if isinstance(value, basestring):
            translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
            if translation.exists():
                translation = translation[0].short_name
            else:
                translation = key
    
            alias_type = Attributes.objects.get(Q(active=True), Q(type='alias_type'), Q(identifier=translation) | Q(name=translation) | Q(short_name=translation))
            if not parent.aliases.filter(alias_type__id=alias_type.id).exists():
                new_alias = Alias()        
                new_alias.alias_type = alias_type
                new_alias.alias_value = value
                new_alias.alias_additional = ''
                new_alias.save()
                return new_alias
            else:
                return parent.aliases.get(alias_type__id=alias_type.id)
        else:
            if value['id']!=None and value['id']!='':
                alias = Alias.objects.get(id=value['id'])
            else:
                alias = Alias()
            for field in value.keys():
                if field not in ['id', 'many-to-many']:
                    if field!='alias_type':
                        setattr(alias, field, value[field])
                    else:
                        alias.alias_type = Attributes.objects.get(active=True, type='alias_type', identifier=value[field])
            alias.save()
            return alias

    class Meta:
        ordering = ['alias_value']

class Container(CoreModel):
    name = models.CharField(max_length=1024)
    short_name = models.CharField(max_length=512)
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='container_type_rel', null=True)
    inception_date = models.DateField(null=True)
    closed_date = models.DateField(null=True)
    short_description = models.TextField(null=True, blank=True)
    status = models.ForeignKey(Attributes,limit_choices_to = {'type':'status'}, related_name='container_status_rel', null=True)
    
    many_fields = {}
    
    @classmethod
    def create(cls):
        entity = cls()
        entity.many_fields = {}
        return entity
    
    @staticmethod
    def get_fields():
        return ['name','short_name','type','inception_date','closed_date','status']
    
    def finalize(self):
        active = Attributes.objects.get(active=True, type='status', identifier='STATUS_ACTIVE')
        self.status = active
        self.save()
        # TODO Loop on many to many only
        for field_name in self._meta.get_all_field_names():
            try:
                if self._meta.get_field(field_name).get_internal_type()=='ManyToManyField':
                    if self.many_fields.has_key(field_name):
                        values = list(self.many_fields[field_name])
                        setattr(self, field_name, values)
            except FieldDoesNotExist:
                None
        self.save()
        
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name', 'short_description']
        
    def clean(self):
        self.many_fields = {}
        for field_name in self._meta.get_all_field_names():
            try:
                if self._meta.get_field(field_name).get_internal_type()=='ForeignKey':
                    setattr(self, field_name, None)
                elif self._meta.get_field(field_name).get_internal_type()=='ManyToManyField':
                    setattr(self, field_name, [])
            except FieldDoesNotExist:
                None

    def clone(self, _class):
        _clone = _class()
        [setattr(_clone, fld.name, getattr(self, fld.name)) for fld in self._meta.fields if fld.name != self._meta.pk and fld.name.find('_ptr')==-1]
        _clone.id = None
        _clone.pk = None
        return _clone

    class Meta:
        ordering = ['name']

class MailCampaignContainer(Container):
    owner = models.ForeignKey(User, related_name='mail_campaign_owner_rel')
    description = models.TextField(null=True, blank=True)
    external_id = models.CharField(max_length=512)
    imported = models.BooleanField(default=False)
    clicked_count = models.IntegerField(default=0)
    opened_count = models.IntegerField(default=0)
    submitted_count = models.IntegerField(default=0)
    unsubscribed_count = models.IntegerField(default=0)
    bounced_count = models.IntegerField(default=0)
    complained_count = models.IntegerField(default=0)
    dropped_count = models.IntegerField(default=0)
    delivered_count = models.IntegerField(default=0)

class Universe(Container):
    public = models.BooleanField()
    members = models.ManyToManyField("Container", related_name='universe_financial_rel')
    owner = models.ForeignKey(User, related_name='universe_owner_rel')
    description = models.TextField(null=True, blank=True)

class FinancialContainer(Container):
    currency = models.ForeignKey(Attributes, limit_choices_to = {'type':'currency'}, related_name='container_currency_rel', null=True)
    owner_role = models.ForeignKey(Attributes, limit_choices_to={'type':'third_party_role'}, related_name='owner_role_rel', null=True)
    owner = models.ForeignKey(Container, limit_choices_to = {'type__identifier':'third_party'}, related_name='container_owner_rel', null=True)
    
    aliases = models.ManyToManyField(Alias, related_name='financial_alias_rel')
    
    associated_companies = models.ManyToManyField("RelatedCompany")
    associated_thirds = models.ManyToManyField("RelatedThird")
    
    frequency = models.ForeignKey(Attributes, limit_choices_to={'type':'frequency'}, related_name='financial_frequency_rel', null=True)
    frequency_reference = models.ForeignKey(Attributes, limit_choices_to={'type':'date_reference'}, related_name='financial_date_reference_rel', null=True)
    
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name', 'short_description']
    
    def update_alias(self, alias_type, value, additional = '', append = False):
        wrk_alias = self.aliases.filter(alias_type__id=alias_type.id)
        if wrk_alias.exists() and not append:
            wrk_alias = wrk_alias[0]
            wrk_alias.alias_value = value
            wrk_alias.alias_additional = additional
            wrk_alias.save()
        else:
            wrk_alias = Alias()
            wrk_alias.alias_type = alias_type
            wrk_alias.alias_value = value
            wrk_alias.alias_additional = additional
            wrk_alias.save()
            self.aliases.add(wrk_alias)
            self.save()
    
    @staticmethod
    def get_fields():
        return Container.get_fields() + ['currency','owner_role','owner','aliases', 'frequency', 'frequency_reference', 'associated_thirds', 'associated_companies']
    
    def get_currency_in_name(self):
        currencies = Attributes.objects.filter(type='currency').values_list('short_name', flat=True)
        for currency in currencies:
            if str(self.name).find(' ' + currency)>0:
                return Attributes.objects.get(type='currency', short_name=currency)
        return None
    
class FinancialOperation(CoreModel):
    name = models.CharField(max_length=512)
    short_name = models.CharField(max_length=128)
    status = models.ForeignKey(Attributes,limit_choices_to = {'type':'operation_status'}, related_name='operation_status_rel', null=True)
    operation_type = models.ForeignKey(Attributes,limit_choices_to = {'type':'operation_type'}, related_name='operation_type_rel', null=True)
    creation_date = models.DateTimeField(default=datetime.datetime.today())
    operator = models.ForeignKey(User, related_name='operator_rel', null=True)
    validator = models.ForeignKey(User, related_name='validator_rel', null=True)
    source = models.ForeignKey('AccountContainer', related_name='source_rel', null=True)
    target = models.ForeignKey('FinancialContainer', related_name='target_rel', null=True)
    spot = models.FloatField(null=True)
    repository = models.ForeignKey('AccountContainer', related_name='repository_rel', null=True)
    quantity = models.FloatField(null=True)
    amount = models.FloatField(null=True)
    price = models.FloatField(null=True)
    operation_date = models.DateTimeField(default=datetime.datetime.today())
    operation_pnl = models.FloatField(null=True)
    value_date = models.DateTimeField(default=datetime.datetime.today())
    termination_date = models.DateTimeField(default=datetime.datetime.today())
    associated_operation = models.ForeignKey('FinancialOperation', related_name='parent_operation_rel', null=True)
    
    @staticmethod
    def get_fields():
        return ['name','short_name','status','operation_type','creation_date','operator','validator','source','target','spot','repository','quantity','amount','price','operation_date','operation_pnl','value_date','termination_date','associated_operation']
    
    class Meta:
        ordering = ['name']

class ThirdPartyContainer(Container):
    addresses = models.ManyToManyField(Address)
    emails = models.ManyToManyField(Email)
    phones = models.ManyToManyField(Phone)

    @staticmethod
    def get_fields():
        return Container.get_fields() + ['addresses','emails','phones']

class PersonContainer(ThirdPartyContainer):
    first_name = models.CharField(max_length=128)
    last_name = models.CharField(max_length=128)
    birth_date = models.DateField(null=True)
    
    @staticmethod
    def get_fields():
        return ThirdPartyContainer.get_fields() + ['first_name','last_name','birth_date']
    
    def get_short_json(self):
        return {'id': self.id, 'first_name': self.first_name, 'last_name': self.last_name}
    
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name', 'short_description', 'first_name', 'last_name']

class CompanySubsidiary(CoreModel):
    company = models.ForeignKey('CompanyContainer', null=True)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'company_subsidiary_role'}, related_name='company_subsidiary_role_rel', null=True)

    def get_identifier(self):
        return 'id'
    
    @staticmethod
    def get_fields():
        return ['company','role']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return universe.models.CompanyContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['company.name', 'company.inception_date', 'company.status.name']
        elif rendition_width=='small':
            return ['company.name']

class CompanyMember(CoreModel):
    person = models.ForeignKey(PersonContainer, null=True)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'company_member_role'}, related_name='company_member_role_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['person','role']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return PersonContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['person.last_name', 'person.first_name', 'person.birth_date']
        elif rendition_width=='small':
            return ['person.last_name', 'person.first_name']

class CompanyContainer(ThirdPartyContainer):
    members = models.ManyToManyField(CompanyMember)
    subsidiary = models.ManyToManyField(CompanySubsidiary)
    
    @staticmethod
    def get_fields():
        return ThirdPartyContainer.get_fields() + ['members','subsidiary']

    def get_short_json(self):
        data_provider = Attributes.objects.get(identifier='SCR_DP', active=True)
        is_provider = RelatedCompany.objects.filter(company=self, role=data_provider).exists()

        return {'id': self.id, 'name': self.name, 'short_name': self.short_name, 'provider': is_provider}
    
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name']

class AccountContainer(FinancialContainer):
    account_type = models.ForeignKey(Attributes, limit_choices_to={'type':'account_type'}, related_name='account_type_rel', null=True)
    bank = models.ForeignKey(CompanyContainer, related_name='account_bank_rel', null=True)
    
    @staticmethod
    def get_fields():
        return FinancialContainer.get_fields() + ['account_type','bank']
    
    @staticmethod
    def get_filtering_field():
        return 'account_type'
    
    @staticmethod
    def get_querying_class():
        return CompanyContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['account_type.name', 'currency.short_name', 'bank.name']
        elif rendition_width=='small':
            return ['account_type.name', 'currency.short_name']
    
class RelatedCompany(CoreModel):
    company = models.ForeignKey(CompanyContainer, null=True)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'security_company_role'}, related_name='security_company_role_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['company','role']

    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['company.name','role.name','company.status.name']
        elif rendition_width=='small':
            return ['company.name','role.name']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return CompanyContainer
    
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent!='web':
            translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
            if translation.exists():
                translation = translation[0].short_name
            else:
                translation = key
            working_value = value
        else:
            translation = value['role']
            working_value = value['company']
            
        company = CompanyContainer.objects.filter(name=working_value)
        if company.exists():
            company = company[0]
        else:
            company = CompanyContainer()
            company.name = working_value
            company.short_name = 'Company...'
            tbv = Attributes.objects.get(active=True, type='status', identifier='STATUS_TO_BE_VALIDATED')
            company.status = tbv
            company.type = Attributes.objects.get(active=True, type='container_type', identifier='CONT_COMPANY')
            company.save()
        role = Attributes.objects.get(Q(active=True), Q(type='security_company_role'), Q(identifier=translation) | Q(name=translation) | Q(short_name=translation))
        relation = RelatedCompany.objects.filter(company__id=company.id, role__id=role.id)
        if relation.exists():
            return relation[0]
        else:
            new_relation = RelatedCompany()        
            new_relation.role = role
            new_relation.company = company
            new_relation.save()
            return new_relation


class RelatedThird(CoreModel):
    third = models.ForeignKey(PersonContainer)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'security_third_role'}, related_name='security_third_role_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['third','role']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return universe.models.PersonContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['third.name','role.name','third.status.name']
        elif rendition_width=='small':
            return ['third.name','role.name']
    
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
        if translation.exists():
            translation = translation[0].short_name
        else:
            translation = key
        third = PersonContainer.objects.filter(name=value)
        if third.exists():
            third = third[0]
        else:
            third = PersonContainer()
            third.name = value
            third.short_name = 'Company...'
            tbv = Attributes.objects.get(active=True, type='status', identifier='STATUS_TO_BE_VALIDATED')
            third.status = tbv
            third.type = Attributes.objects.get(active=True, type='container_type', identifier='CONT_PERSON')
            third.save()
            
        new_relation = RelatedThird()        
        new_relation.role = Attributes.objects.get(Q(active=True), Q(type='security_third_role'), Q(identifier=translation) | Q(name=translation) | Q(short_name=translation))
        new_relation.third = third
        new_relation.save()
        return new_relation
    
class PortfolioContainer(FinancialContainer):
    accounts = models.ManyToManyField('AccountContainer', related_name='portfolio_accounts_rel')
    
    @staticmethod
    def get_fields():
        return FinancialContainer.get_fields() + ['accounts']
    
    def get_short_json(self):
        return {'id': self.id, 'name': self.name, 'short_name': self.short_name, 'currency': self.currency.short_name if self.currency!=None else '', 'manager':'', 'bank':'', 'aum':0.0}

class BloombergDataContainerMapping(CoreModel):
    name = models.CharField(max_length=64)
    short_name = models.ForeignKey(BloombergField, limit_choices_to={'get_history':False}, related_name='bloomberg_field_data_rel')
    container = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='bloomberg_container_type_rel')
    model_link = models.CharField(max_length=128, null=True)
    model_visible = models.BooleanField(default=False)
    active = models.BooleanField(default=True)

    @staticmethod
    def get_fields():
        return ['name','field','container','model_link','model_visible','active']

class BloombergTrackContainerMapping(CoreModel):
    name = models.ForeignKey(Attributes, limit_choices_to={'type':'numeric_type'}, related_name='bb_numeric_type_rel')
    short_name = models.ForeignKey(BloombergField, limit_choices_to={'get_history':True}, related_name='bloomberg_field_track_rel')
    container = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='bloomberg_container_track_rel')
    active = models.BooleanField(default=True)
    
    @staticmethod
    def get_fields():
        return ['name', 'field', 'container', 'active']


class SecurityContainer(FinancialContainer):
    security_type = models.ForeignKey(Attributes, limit_choices_to={'type':'security_type'}, related_name='security_type_rel', null=True)
    country = models.ForeignKey(Attributes, limit_choices_to={'type':'country_iso2'}, related_name='security_country_rel', null=True)
    region = models.CharField(max_length=128, null=True, blank=True)
    market_sector = models.CharField(max_length=128, null=True, blank=True)

    parent_security = models.ForeignKey('SecurityContainer', related_name='parent_security_rel', null=True)
    
    attached_account = models.ForeignKey('AccountContainer', related_name='financials_account_rel', null=True)
    
    @staticmethod
    def get_fields():
        return FinancialContainer.get_fields() + ['security_type','country','region','market_sector', 'parent_security','attached_account']
    
    def get_short_json(self):
        isin = self.aliases.filter(alias_type__short_name='ISIN')
        bloomberg = self.aliases.filter(alias_type__short_name='BLOOMBERG')
        if isin.exists():
            isin = isin[0].alias_value
        else:
            isin = ''
        if bloomberg.exists():
            bloomberg = bloomberg[0].alias_value
        else:
            bloomberg = ''
        
        if self.currency!=None:
            currency = self.currency.short_name
        else:
            currency = ''
        
        return {'id': self.id, 'name': self.name, 'short_name': self.short_name, 'currency': currency, 'isin': isin, 'bloomberg': bloomberg}
    
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name', 'short_description']
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['name', 'type.name', 'currency.short_name', 'inception_date', 'closed_date', 'status.name']
        elif rendition_width=='small':
            return ['name', 'type.name', 'currency.short_name']

class BacktestContainer(FinancialContainer):
    universe = models.ForeignKey(Universe, related_name='backtest_universe_rel')
    public = models.BooleanField()
    publisher = models.ForeignKey(User, related_name='backtest_publisher_rel')
    reweight = models.BooleanField()
    organic_date = models.DateField(null=True)
    hedging = models.ForeignKey(Attributes, limit_choices_to={'type':'hedging_method'}, related_name='backtest_hedging_method_rel')
    fees = models.ForeignKey(Attributes, limit_choices_to={'type':'fees_scheme'}, related_name='backtest_fees_scheme_rel')
    leverage = models.ForeignKey(Attributes, limit_choices_to={'type':'leverage'}, related_name='backtest_leverage_rel')
    leverage_level = models.FloatField(default=0.0)
    history = models.ForeignKey(Attributes, limit_choices_to={'type':'history_completion'}, related_name='backtest_history_completion_rel')
    initial_aum = models.FloatField()
    
    @staticmethod
    def get_fields():
        return FinancialContainer.get_fields() + ['universe','public','publisher','from_date','to_date','reweight','organic_date','hedging','fees','leverage','leverage_level','history','initial_aum']

class TrackContainer(CoreModel):
    container = models.ForeignKey(ContentType)
    effective_container_id = models.PositiveIntegerField()
    effective_container = generic.GenericForeignKey('container', 'effective_container_id')
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'numeric_type'}, related_name='numeric_type_rel', null=True)
    frequency = models.ForeignKey(Attributes, limit_choices_to={'type':'frequency'}, related_name='numeric_frequency_rel')
    frequency_reference = models.ForeignKey(Attributes, limit_choices_to={'type':'date_reference'}, related_name='numeric_date_reference_rel', null=True)
    quality = models.ForeignKey(Attributes, limit_choices_to={'type':'numeric_quality'}, related_name='numeric_quality_rel', null=True)
    status = models.ForeignKey(Attributes, limit_choices_to={'type':'numeric_status'}, related_name='numeric_status_rel', null=True)
    source = models.ForeignKey(CompanyContainer, related_name='data_source_rel', null=True)
    start_date = DateTimeField(null=True)
    end_date = DateTimeField(null=True)
    
    def get_identifier(self):
        return 'id'
    
    @staticmethod
    def get_fields():
        return ['effective_container','type','quality','source','start_date','end_date']
