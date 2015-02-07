import datetime
from utilities.external_content import get_positions
import logging
import sys
from universe.models import Universe, Alias, Attributes, SecurityContainer,\
    TrackContainer, CompanyContainer, Address, PersonContainer, CompanyMember,\
    Phone, Email
from utilities.track_content import get_track_content
from seq_common.utils import dates
import traceback
import os
from finale.settings import WORKING_PATH
from openpyxl.workbook import Workbook
from external.sequoia_data import ROLE_MAPS_TEMPLATE, SEQUOIA_STYLES
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.reader.excel import load_workbook
from django.db.models import Q
from django.contrib.auth.models import User

LOGGER = logging.getLogger(__name__)

def import_crm_file(file_name):
    LOGGER.info("Loading data from " + file_name)
    workbook = load_workbook(file_name)
    sheet = workbook.get_sheet_by_name('Investment managers prospect 20')
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value.strip() if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    LOGGER.info('Rows:' + str(sheet.get_highest_row()))
    LOGGER.info('Columns:' + str(sheet.get_highest_column()))
    
    company_type = Attributes.objects.get( active=True, type='container_type', identifier='CONT_COMPANY')
    person_type = Attributes.objects.get( active=True, type='container_type', identifier='CONT_PERSON')
    active = Attributes.objects.get(active=True, type='status', identifier='STATUS_ACTIVE')
    closed = Attributes.objects.get(active=True, type='status', identifier='STATUS_CLOSED')
    main_address = Attributes.objects.get(active=True, type='address_type', identifier='ADR_MAIN_OFFICE')
    secondary_address = Attributes.objects.get(active=True, type='address_type', identifier='ADR_SUBSI_OFFICE')
    web_address = Attributes.objects.get(active=True, type='address_type', identifier='ADR_WEB')
    mobile_work_phone = Attributes.objects.get(active=True, type='phone_type', identifier='PHONE_MOB_WORK')
    land_work_phone = Attributes.objects.get(active=True, type='phone_type', identifier='PHONE_LAND_WORK')
    email_work = Attributes.objects.get(active=True, type='email_type', identifier='EMAIL_WORK')
    
    row_index += 1
    
    company_name_index = header.index('Company') + 1
    description_index = header.index('Company Info') + 1
    active_index = header.index('Dead/Alive') + 1
    address_line_1_index = header.index('Address') + 1
    address_zip_index = header.index('Zip Code') + 1
    address_city_index = header.index('City') + 1
    address_country_index = header.index('Country') + 1
    website_index = header.index('Internet site') + 1
    
    person_first_index = header.index('First name') + 1
    person_last_index = header.index('Family name') + 1
    person_title_index = header.index('Title') + 1
    person_office_phone_index = header.index('Office phone') + 1
    person_mobile_phone_index = header.index('Mobile') + 1
    person_email_index = header.index('Email') + 1
    
    universe = Universe()
    universe.name = 'CRM Universe ' + str(datetime.datetime.now())
    universe.short_name = str(datetime.datetime.now())
    universe.type = Attributes.objects.get(identifier='CONT_UNIVERSE')
    universe.inception_date = datetime.date.today()
    universe.closed_date = None
    universe.status = Attributes.objects.get(identifier='STATUS_ACTIVE')
    universe.public = True
    universe.owner = User.objects.get(id=1)
    universe.description = "This universe contains all CRM related companies as of " + str(datetime.datetime.now()) + "."
    universe.save()
    
    treated_company = []
    
    while row_index<=sheet.get_highest_row():
        LOGGER.info("Working on row:" + str(row_index))
        if sheet.cell(row = row_index, column=company_name_index).value==None or sheet.cell(row = row_index, column=company_name_index).value=='':
            row_index += 1
            continue
        company_name = sheet.cell(row = row_index, column=company_name_index).value.strip()
        company = CompanyContainer.objects.filter(Q(name=company_name) | Q(short_name=company_name))
        
        is_active = sheet.cell(row = row_index, column=active_index).value.strip().upper()!='DEAD' if sheet.cell(row = row_index, column=active_index).value!=None else True

        if company.exists():
            company = company[0]
            LOGGER.info("Company " + company_name + " already exists, updating.")
            if company_name not in treated_company:
                for phone in company.phones.all():
                    company.phones.remove(phone)
                    company.save()
                    phone.delete()
                for address in company.addresses.all():
                    company.addresses.remove(address)
                    company.save()
                    address.delete()
                for member in company.members.all():
                    company.members.remove(member)
                    company.save()
                    member.delete()
        else:
            LOGGER.info("Company " + company_name + " is being created.")
            company = CompanyContainer()
            company.type = company_type
            company.name = company_name
            company.short_name =  company_name
            company.status = active if is_active else closed
            company.save()
        
        company.short_description = sheet.cell(row = row_index, column=description_index).value.strip() if sheet.cell(row = row_index, column=description_index).value!=None else None
        
        country_name = sheet.cell(row = row_index, column=address_country_index).value.strip().upper() if sheet.cell(row = row_index, column=address_country_index).value!=None else ''
        country_attribute = Attributes.objects.filter(active=True, type='country_iso2', name__iexact=country_name)
        if country_attribute.exists():
            country_attribute = country_attribute[0]
        else:
            country_attribute = None
        if country_attribute!=None \
            or sheet.cell(row = row_index, column=address_line_1_index).value!=None \
            or sheet.cell(row = row_index, column=address_zip_index).value!=None \
            or sheet.cell(row = row_index, column=address_city_index).value!=None:
            
            line_1_value = sheet.cell(row = row_index, column=address_line_1_index).value.strip().upper() if sheet.cell(row = row_index, column=address_line_1_index).value!=None else None
            zip_code_value = str(sheet.cell(row = row_index, column=address_zip_index).value).strip().upper() if sheet.cell(row = row_index, column=address_zip_index).value!=None else None
            city_value = sheet.cell(row = row_index, column=address_city_index).value.strip().upper() if sheet.cell(row = row_index, column=address_city_index).value!=None else None
            addresses = company.addresses.filter(line_1=line_1_value,
                                         zip_code=zip_code_value,
                                         city=city_value,
                                         country=country_attribute)
            if not addresses.exists():
                company.save()
                address = Address()
                address.address_type = main_address if len(addresses)==0 else secondary_address
                address.line_1 = line_1_value
                address.line_2 = None
                address.zip_code = zip_code_value
                address.city = city_value
                address.country = country_attribute
                address.save()
                company.addresses.add(address)
                company.save()
        if sheet.cell(row = row_index, column=website_index).value!=None and sheet.cell(row = row_index, column=website_index).value!='':
            website_url = str(sheet.cell(row = row_index, column=website_index).value).strip()
            website = company.addresses.filter(address_type=web_address, line_1=website_url)
            if not website.exists() and website_url!='' and website_url!=None:
                website = Address()
                website.address_type = web_address
                website.line_1 = website_url
                website.line_2 = None
                website.zip_code = None
                website.city = None
                website.country = None
                website.save()
                company.addresses.add(website)
                company.save()
        
        person_first = sheet.cell(row = row_index, column=person_first_index).value.strip() if sheet.cell(row = row_index, column=person_first_index).value!=None and sheet.cell(row = row_index, column=person_first_index).value!='' else None
        person_last = sheet.cell(row = row_index, column=person_last_index).value.strip().upper() if sheet.cell(row = row_index, column=person_last_index).value!=None and sheet.cell(row = row_index, column=person_last_index).value!='' else None
        
        if person_first!=None and person_last!=None:
            persons = PersonContainer.objects.filter(last_name=person_last, first_name=person_first)
            if not persons.exists():
                LOGGER.info("Person " + person_last + ' ' + person_first + " is being created.")
                person = PersonContainer()
                person.first_name = person_first
                person.last_name = person_last
                person.birth_date = None
                person.name = person_last + ' ' + person_first
                person.type = person_type
                person.status = active if is_active else closed
                person.save()
            else:
                LOGGER.info("Person " + person_last + ' ' + person_first + " already exists, updating.")
                person = persons[0]
                for phone in person.phones.all():
                    person.phones.remove(phone)
                    person.save()
                    phone.delete()
                for email in person.emails.all():
                    person.emails.remove(email)
                    person.save()
                    email.delete()
            if sheet.cell(row = row_index, column=person_title_index).value!=None and sheet.cell(row = row_index, column=person_title_index).value!='':
                titles = sheet.cell(row = row_index, column=person_title_index).value.split('/')
                for title in titles:
                    current_title = Attributes.objects.filter(Q(short_name__iexact=title.strip()) | Q(name__iexact=title.strip()), Q(type='company_member_role'), Q(active=True))
                    if current_title.exists():
                        LOGGER.info("Person " + person_last + ' ' + person_first + " is assigned as " + title + ".")
                        member = CompanyMember()
                        member.person = person
                        member.role = current_title[0]
                        member.save()
                        company.members.add(member)
                    else:
                        LOGGER.warn("Title [" + title + "] doesn't exists please correct your data or add that role to the system." )
            else:
                LOGGER.warn("Person " + person_last + ' ' + person_first + " has no title within the company.")
                
            if sheet.cell(row = row_index, column=person_mobile_phone_index).value!=None and sheet.cell(row = row_index, column=person_mobile_phone_index).value!='':
                mobile = Phone()
                mobile.phone_number = sheet.cell(row = row_index, column=person_mobile_phone_index).value.strip()
                mobile.line_type = mobile_work_phone
                mobile.save()
                person.phones.add(mobile)
                person.save()
            if sheet.cell(row = row_index, column=person_office_phone_index).value!=None and sheet.cell(row = row_index, column=person_office_phone_index).value!='':
                land = Phone()
                land.phone_number = sheet.cell(row = row_index, column=person_office_phone_index).value.strip()
                land.line_type = land_work_phone
                land.save()
                person.phones.add(land)
                person.save()
            if sheet.cell(row = row_index, column=person_email_index).value!=None and sheet.cell(row = row_index, column=person_email_index).value!='':
                email = Email()
                email.email_address = sheet.cell(row = row_index, column=person_email_index).value.strip()
                email.address_type = email_work
                email.save()
                person.emails.add(email)
                person.save()
        company.save()
        universe.members.add(company)
        if company_name not in treated_company:
            treated_company.append(company_name)
        row_index += 1
    universe.save()


def define_xslx_header(sheet, headers):
    row_index = 1
    col_index = 1
    for row in headers:
        for column in row:
            sheet.cell(row=row_index, column=col_index).value = column
            col_index += 1
        row_index += 1
        col_index = 1
    return row_index

def get_name_from_identifier(identifier):
    return Attributes.objects.get(identifier=identifier, active=True).name

def get_details(data, target, fees_type, index):
    if index<len(data[fees_type][target]):
        print data[fees_type][target][index]
        return {'bud': get_name_from_identifier(data[fees_type][target][index]['bud']), 'rate': data[fees_type][target][index]['rate'] / 100.0}
    else:
        return {'bud': None, 'rate': None}

def execute_row_command(ws, row, row_index, container, data_map, query_date):
    hide = False
    if row!=None:
        LOGGER.debug("Row:" + str(row_index) + ' - ' + str(row))
        for col_index in row.keys():
            if row[col_index].has_key('text'):
                value = row[col_index]['text']
                if not isinstance(row[col_index]['text'], basestring):
                    # value = [ row[col_index]['text'][i] if i%2==1 else all_formats[row[col_index]['text'][i]] for i in range(0, len(row[col_index]['text']))]
                    value = 'TO BE IMPLEMENTED'
            elif row[col_index].has_key('value'):
                try:
                    value = eval(row[col_index]['value'])
                except:
                    if row[col_index].has_key('default'):
                        value = row[col_index]['default']
                    else:
                        value = ''
                if row[col_index].has_key('field'):
                    if value!=None and isinstance(value, dict) and value[row[col_index]['field']]!=None:
                        value = value[row[col_index]['field']]
                    else:
                        value = '' if row[col_index]['default']==None else row[col_index]['default']
                        print "->" + str(value)
                        hide = hide or (row[col_index]['on_default_hide'] if row[col_index].has_key('on_default_hide') else False)
            elif row[col_index].has_key('formula'):
                # Same as value for the time being, may change in the future
                value = row[col_index]['formula']
            ws.cell(row=row_index, column=col_index).value = value
            if row[col_index].has_key('merge'):
                row_span = row[col_index]['merge']['row_span'] if row[col_index]['merge'].has_key('row_span') else 0
                ws.merge_cells(start_row=row_index, start_column=row[col_index]['merge']['start'], end_row=row_index + row_span,end_column=row[col_index]['merge']['end'])
                for merge_row_index in range(row_index, row_index + row_span + 1):
                    for merge_col_index in range(row[col_index]['merge']['start'], row[col_index]['merge']['end'] + 1):
                        ws.cell(row=merge_row_index, column=merge_col_index).style = SEQUOIA_STYLES['SWM'][row[col_index]['format']]
            else:
                ws.cell(row=row_index, column=col_index).style = SEQUOIA_STYLES['SWM'][row[col_index]['format']]
                # ws.write(row_index, col_index, value, all_formats[row[col_index]['format']])
    else:
        ws.cell(row=row_index, column=1).value = ''
    return hide

def export_map(container, data, workbook = None):
    if workbook==None:
        workbook = Workbook()
        sheet = workbook.active
    else:
        sheet = workbook.create_sheet()
    sheet.title = container.short_name
    xlsx_rows_setup = ROLE_MAPS_TEMPLATE['SWM']
    row_index = 1
    for row in xlsx_rows_setup:
        hide = execute_row_command(sheet, row, row_index, container, data, datetime.date.today() )
        print str(row_index) + ' - ' + str(hide)
        if row_index==1:
            sheet.row_dimensions[row_index].height = 33.75
        elif row_index==2:
            sheet.row_dimensions[row_index].height = 32.25
        elif row_index>74:
            sheet.row_dimensions[row_index].height = 19.5
        else:
            if hide:
                sheet.row_dimensions[row_index] = RowDimension(hidden=True)
            else:
                sheet.row_dimensions[row_index].height=27.0
        row_index = row_index + 1
    for col in range(1,10):
        sheet.cell(row=row_index, column=col).value = ''
    sheet.column_dimensions['A'].width = 35.71
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 47.71
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 8.43
    sheet.column_dimensions['F'].width = 5.29
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 8.43
    sheet.column_dimensions['I'].width = 8.43
    return workbook

def export(parameters):
    working_date = parameters['workingDate']
    export_type = parameters['exportType']
    
    sequoia_universe = Universe.objects.get(name='Sequoia Products')
    guardian_alias = Attributes.objects.get(type='alias_type', short_name='GUARDIAN')
    bloomberg_alias = Attributes.objects.get(identifier='ALIAS_BLOOMBERG')
    
    nav_value = Attributes.objects.get(identifier='NUM_TYPE_NAV', active=True)
    final_status = Attributes.objects.get(identifier='NUM_STATUS_FINAL', active=True)
    official_type = Attributes.objects.get(identifier='PRICE_TYPE_OFFICIAL', active=True)
    monthly = Attributes.objects.get(identifier='FREQ_MONTHLY', active=True)
    computing_company = CompanyContainer.objects.get(name='FinaLE Engine')
    
    if not isinstance(working_date, basestring):
        working_date = working_date[0]
    end_date = datetime.datetime.strptime(working_date,'%Y-%m-%d')
    begin_date = dates.GetStartOfMonth(end_date)
    all_portfolios = get_positions('guardian', end_date)
    
    all_products = {}
    all_groups = {}
    all_navs = {}
    all_spots = {'CHF': 1.0}
    
    for portfolio_code in all_portfolios['values']:
        portfolio_name = portfolio_code.replace('___','.')
        LOGGER.info("Working on " + portfolio_name)
        
        all_positions = all_portfolios['values'][portfolio_code]
        for position in all_positions:
            group = position['cod_gru']
            printable_name = position['des_tit'].encode(sys.stdout.encoding, errors='replace')
            target_currency = position['cod_div_tit']

            if not all_groups.has_key(group):
                all_groups[group] = {}
                
            if not all_groups[group].has_key(portfolio_name):
                all_groups[group][portfolio_name] = {'AUM': 0.0}
                
            security = SecurityContainer.objects.filter(aliases__alias_type=guardian_alias, aliases__alias_value=position['cod_tit'])
            if security.exists():
                security = security[0]
                if target_currency==None and target_currency.currency!=None:
                    target_currency = security.currency.short_name
            else:
                security = None
            print "SECURITY:" + printable_name               
            if security!=None:
                sequoia = sequoia_universe.members.filter(id=security.id).exists()
                if sequoia:
                    LOGGER.info("\tFound " + printable_name)
                    if not all_groups[group][portfolio_name].has_key('SEQUOIA'):
                        all_groups[group][portfolio_name]['SEQUOIA'] = {}
                    if not all_products.has_key(security.name):
                        all_products[security.name] = {}
                if not all_navs.has_key(position['des_tit']):
                    try:
                        monthly_track = TrackContainer.objects.get(effective_container_id=security.id, type__id=nav_value.id,quality__id=official_type.id, source__id=computing_company.id, frequency__id=monthly.id, status__id=final_status.id)
                        monthlies = {token['date']:token['value'] for token in get_track_content(monthly_track)}
                        backdated_day = end_date
                        while not monthlies.has_key(backdated_day):
                            backdated_day = dates.AddDay(backdated_day, -1)
                        nav = monthlies[backdated_day]
                        all_navs[position['des_tit']] = nav
                    except:
                        LOGGER.warn("\tNo price nor track for " + printable_name)
                        all_navs[position['des_tit']] = 1.0
                if not all_spots.has_key(target_currency):
                    spot_container = SecurityContainer.objects.filter(aliases__alias_type=bloomberg_alias, aliases__alias_value= target_currency + 'CHF Curncy')
                    if spot_container.exists():
                        spot_container = spot_container[0]
                        try:
                            monthly_track = TrackContainer.objects.get(effective_container_id=spot_container.id, type__id=nav_value.id,quality__id=official_type.id, source__id=computing_company.id, frequency__id=monthly.id, status__id=final_status.id)
                            monthlies = {token['date']:token['value'] for token in get_track_content(monthly_track)}
                            backdated_day = end_date
                            while not monthlies.has_key(backdated_day):
                                backdated_day = dates.AddDay(backdated_day, -1)
                            spot = monthlies[backdated_day]
                            all_spots[target_currency] = spot
                        except:
                            traceback.print_exc()
                            LOGGER.warn("\tNo price nor track for " + target_currency + "/CHF SPOT")
                            all_spots[target_currency] = 1.0
                    else:
                        LOGGER.warn("\tNo security for " + target_currency + "/CHF SPOT")
                        all_spots[target_currency] = 1.0
                chf_value = position['qta'] * all_navs[position['des_tit']] * all_spots[target_currency]
                if sequoia:
                    all_products[security.name][portfolio_name] = chf_value
                    all_groups[group][portfolio_name]['SEQUOIA'][position['des_tit']] = chf_value
                all_groups[group][portfolio_name]['AUM'] += chf_value
            else:
                if not all_spots.has_key(target_currency):
                    spot_container = SecurityContainer.objects.filter(aliases__alias_type=bloomberg_alias, aliases__alias_value= target_currency + 'CHF Curncy')
                    if spot_container.exists():
                        spot_container = spot_container[0]
                        try:
                            monthly_track = TrackContainer.objects.get(effective_container_id=spot_container.id, type__id=nav_value.id,quality__id=official_type.id, source__id=computing_company.id, frequency__id=monthly.id, status__id=final_status.id)
                            monthlies = {token['date']:token['value'] for token in get_track_content(monthly_track)}
                            backdated_day = end_date
                            while not monthlies.has_key(backdated_day):
                                backdated_day = dates.AddDay(backdated_day, -1)
                            spot = monthlies[backdated_day]
                            all_spots[target_currency] = spot
                        except:
                            traceback.print_exc()
                            LOGGER.warn("\tNo price nor track for " + target_currency + "/CHF SPOT")
                            all_spots[target_currency] = 1.0
                    else:
                        LOGGER.warn("\tNo security for " + target_currency + "/CHF SPOT")
                        all_spots[target_currency] = 1.0
                chf_value = position['qta'] * all_spots[target_currency]
                all_groups[group][portfolio_name]['AUM'] += chf_value
    # OUTPUT of SWM file
    out_file_name = os.path.join(WORKING_PATH,'_SWM_' + str(working_date) + '.xlsx')
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Import"
    headers = [['','','','Allocation',''],['Code Fond / Banque',u'Date d\xe9but','Date fin','Code Client','Montant CHF'],]
    row_index = define_xslx_header(ws, headers)
    for product in sorted(all_products.keys()):
        first = True
        for client in all_products[product].keys():
            if not first:
                ws.cell(row=row_index, column=1).value = ''
            else:
                ws.cell(row=row_index, column=1).value = product
            ws.cell(row=row_index, column=2).value = begin_date
            ws.cell(row=row_index, column=3).value = working_date
            ws.cell(row=row_index, column=4).value = client if client!='NEW FRONTIER' else 'NF'
            ws.cell(row=row_index, column=5).value = all_products[product][client]
            row_index += 1
            first = False
    wb.save(out_file_name)